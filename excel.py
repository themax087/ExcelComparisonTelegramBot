from functools import lru_cache

from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from xlrd import open_workbook
from xlutils.copy import copy
from xlwt import XFStyle


class XlsFile:
    def __init__(self, filename):
        self.filename = filename
        self.book = open_workbook(filename, formatting_info=True)
        self.sheets = {}

    @property
    @lru_cache
    def sheets_names(self):
        return set(self.book.sheet_names())

    @lru_cache
    def sheet_by_name(self, name):
        return self.book.sheet_by_name(name)

    @lru_cache
    def rows(self, name):
        return self.sheet_by_name(name).nrows

    @lru_cache
    def cols(self, name):
        return self.sheet_by_name(name).ncols

    @lru_cache
    def cell_value(self, name, row, col):
        sheet = self.sheet_by_name(name)
        return sheet.cell_value(rowx=row, colx=col)

    def copy_with_highlights(self, path, highlights_by_sheet):
        new_wb = copy(self.book)
        for name, highlights in highlights_by_sheet.items():
            old_sheet = self.sheet_by_name(name)
            sheet = new_wb.get_sheet(name)
            for row, col in highlights:
                old_style = self.book.xf_list[old_sheet.cell(row, col).xf_index]
                sheet.write(
                    row, col, self.cell_value(name, row, col),
                    self.copy_style_with_red_back(old_style),
                )
        new_wb.save(path)

    def copy_style_with_red_back(self, rdxf):
        wtxf = XFStyle()

        wtxf.pattern.pattern = 1
        wtxf.pattern.pattern_fore_colour = 10
        wtxf.pattern.pattern_back_colour = 60

        wtxf.num_format_str = self.book.format_map[rdxf.format_key].format_str

        wtf = wtxf.font
        rdf = self.book.font_list[rdxf.font_index]
        wtf.height = rdf.height
        wtf.italic = rdf.italic
        wtf.struck_out = rdf.struck_out
        wtf.outline = rdf.outline
        wtf.shadow = rdf.outline
        wtf.colour_index = rdf.colour_index
        wtf.bold = rdf.bold
        wtf._weight = rdf.weight
        wtf.escapement = rdf.escapement
        wtf.underline = rdf.underline_type
        wtf.family = rdf.family
        wtf.charset = rdf.character_set
        wtf.name = rdf.name

        wtp = wtxf.protection
        rdp = rdxf.protection
        wtp.cell_locked = rdp.cell_locked
        wtp.formula_hidden = rdp.formula_hidden

        wtb = wtxf.borders
        rdb = rdxf.border
        wtb.left = rdb.left_line_style
        wtb.right = rdb.right_line_style
        wtb.top = rdb.top_line_style
        wtb.bottom = rdb.bottom_line_style
        wtb.diag = rdb.diag_line_style
        wtb.left_colour = rdb.left_colour_index
        wtb.right_colour = rdb.right_colour_index
        wtb.top_colour = rdb.top_colour_index
        wtb.bottom_colour = rdb.bottom_colour_index
        wtb.diag_colour = rdb.diag_colour_index
        wtb.need_diag1 = rdb.diag_down
        wtb.need_diag2 = rdb.diag_up

        wta = wtxf.alignment
        rda = rdxf.alignment
        wta.horz = rda.hor_align
        wta.vert = rda.vert_align
        wta.dire = rda.text_direction
        wta.rota = rda.rotation
        wta.wrap = rda.text_wrapped
        wta.shri = rda.shrink_to_fit
        wta.inde = rda.indent_level

        return wtxf


class XlsxFile:
    def __init__(self, filename):
        self.filename = filename
        self.book = load_workbook(filename)
        self.sheets = {}

    @property
    @lru_cache
    def sheets_names(self):
        return set(self.book.sheetnames)

    @lru_cache
    def sheet_by_name(self, name):
        return self.book[name]

    @lru_cache
    def rows(self, name):
        return self.sheet_by_name(name).max_row

    @lru_cache
    def cols(self, name):
        return self.sheet_by_name(name).max_column

    @lru_cache
    def cell_value(self, name, row, col):
        sheet = self.sheet_by_name(name)
        value = sheet.cell(row + 1, col + 1).value
        if value:
            return value
        return ''

    def copy_with_highlights(self, path, highlights_by_sheet):
        red_fill = PatternFill(start_color="FF0000", fill_type="solid")
        for name, highlights in highlights_by_sheet.items():
            for row, col in highlights:
                self.book[name].cell(row + 1, col + 1).fill = red_fill
        self.book.save(path)


def open_excel_file(filename: str):
    if filename.endswith('.xls'):
        return XlsFile(filename)
    elif filename.endswith('.xlsx'):
        return XlsxFile(filename)
    else:
        raise NotImplementedError


def find_diff(current, original):
    highlights_by_sheet = {}
    sheets_names = current.sheets_names.intersection(original.sheets_names)
    for name in sheets_names:
        sheet_highlights = set()

        current_rows = current.rows(name)
        original_rows = original.rows(name)

        current_cols = current.cols(name)
        original_cols = original.cols(name)

        if current_rows > original_rows:
            for row in range(original_rows, current_rows):
                for col in range(current_cols):
                    sheet_highlights.add((row, col))

        if current_cols > original_cols:
            for col in range(original_cols, current_cols):
                for row in range(current_rows):
                    sheet_highlights.add((row, col))

        for row in range(min(current_rows, original_rows)):
            for col in range(min(current_cols, original_cols)):
                if current.cell_value(name, row, col) != original.cell_value(name, row, col):
                    sheet_highlights.add((row, col))

        highlights_by_sheet[name] = sheet_highlights

    return highlights_by_sheet
